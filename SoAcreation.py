# Creating SoA based off of the Master invoices list
"""OUTLIN
create a new file with SoA template with openoyxl
read in master file as csv
make a loop
for each COMPANY NAME and while status==billed :
    if(END OF INVOICE)
        New sheet
    add invoice number, date, total to NEW FILE

SAVE + CLOSE BOTH
***********************************
important cells in SoA :
EoF = A31 (17 cells)
start description = A15
start date = G15
start charge = h15
*************************************
important columns in Master Invoice (STARTS AT ROW 2)
date: B
Invoice number = A
Name = C
Total $ = I
Payment stat= J

"""
import openpyxl
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('SoA Template.xlsx')

ws = wb.active

# open the master list

df = pd.read_excel('MasterSales.xlsx')

df = df.filter(items=['Invoice', 'Date', 'Name', 'Total Due', 'Payment Status'])

name = df[df['Name'] == 'Canex Freight Systems']
unpaid = name[name['Payment Status'] == 'Billed']
# now unpaid is filtered with all the info we need
# need to copy the invoice #, date and total into wb
pluspage = False
rows = len(unpaid.index)

# add extra sheets if there's too many rows
if rows > 17:
    i = 17 // rows
    for x in range(i):
        wb.copy_worksheet(ws)
# just to make sure we're at first worksheet
ws = wb.active

sheet=1
totalSheets = len(wb.sheetnames)
current_row=15
for i in range(rows):
    ws.cell(row=current_row, column=1, value=unpaid.iloc[i]['Invoice'])
    ws.cell(row=current_row, column=7, value=unpaid.iloc[i]['Date'])
    ws.cell(row=current_row, column=8, value=unpaid.iloc[i]['Total Due'])
    current_row += 1
    if current_row>31 and sheet<totalSheets :
        wb.active=sheet
        ws=wb.active
        sheet+=1
        current_row=16 #leave extra space for page total

wb.save("test.xlsx")


